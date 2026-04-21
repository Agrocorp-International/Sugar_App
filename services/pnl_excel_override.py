"""Parse Summary sheet of sugarm2m.xlsm into an override dict for the dashboard PNL table.

Cells are formula-driven, so we need the *cached* values from openpyxl — which means
the workbook must have been saved by Excel with recalculated formulas. openpyxl never
recalculates; the error messages below surface this caveat when cells come back blank.
"""
import hashlib
import logging

import openpyxl

from models.db import db, PnlOverride

logger = logging.getLogger(__name__)

REQUIRED_SHEET = "Summary"
VALID_SLOTS = ("current", "daily", "weekly", "monthly")


def _validate_slot(slot: str) -> str:
    if slot not in VALID_SLOTS:
        raise ValueError(f"invalid slot: {slot!r}; must be one of {VALID_SLOTS}")
    return slot


SUMMARY_CELLS = {
    "alpha_m2m":           "B3",
    "alpha_pnl":           "B4",
    "net_alpha_pnl":       "B5",
    "whites_physical_m2m": "B8",
    "whites_futures_m2m":  "B9",
    "whites_pnl":          "B10",
    "raws_physical_m2m":   "B13",
    "raws_futures_m2m":    "B14",
    "ffa_m2m":             "B15",
    "net_raws_pnl":        "B17",
    "total_pnl":           "B23",
}


def parse_summary(wb) -> dict:
    """Return {dashboard_key: float} for all 11 mapped cells.

    Raises ValueError if the Summary sheet is missing or any required cell is
    non-numeric. The error lists every bad cell and hints at the openpyxl recalc
    caveat, which is the most common cause of blank cached values.
    """
    if REQUIRED_SHEET not in wb.sheetnames:
        head = ", ".join(wb.sheetnames[:10])
        suffix = "..." if len(wb.sheetnames) > 10 else ""
        raise ValueError(
            f"Missing required sheet '{REQUIRED_SHEET}'. Found: {head}{suffix}"
        )

    ws = wb[REQUIRED_SHEET]
    out: dict = {}
    bad: list = []
    for key, addr in SUMMARY_CELLS.items():
        v = ws[addr].value
        if isinstance(v, bool):
            bad.append(f"{addr} ({key}) = {v!r}")
        elif isinstance(v, (int, float)):
            out[key] = float(v)
        else:
            bad.append(f"{addr} ({key}) = {v!r}")

    if bad:
        raise ValueError(
            "Non-numeric or blank cells in Summary sheet: "
            + "; ".join(bad)
            + ". If these are formulas, open the workbook in Excel and Save once "
              "to refresh cached values (openpyxl cannot recalculate formulas)."
        )
    return out


def parse_workbook_from_path(path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return parse_summary(wb)
    finally:
        wb.close()


def parse_workbook_from_file(file_storage) -> dict:
    wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    try:
        return parse_summary(wb)
    finally:
        wb.close()


def sha256_of_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_of_path(path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def current_user_label() -> str:
    import getpass
    import socket
    try:
        user = getpass.getuser()
    except Exception:
        user = "unknown"
    try:
        host = socket.gethostname()
    except Exception:
        host = "unknown-host"
    return f"{user}@{host}"


def get_active_override(slot: str = "current") -> PnlOverride | None:
    _validate_slot(slot)
    return (PnlOverride.query
            .filter_by(slot=slot, is_active=True)
            .order_by(PnlOverride.uploaded_at.desc(), PnlOverride.id.desc())
            .first())


def get_active_overrides_by_slot() -> dict:
    """Return {slot: PnlOverride|None} for all 4 slots in one query."""
    rows = (PnlOverride.query
            .filter_by(is_active=True)
            .order_by(PnlOverride.uploaded_at.desc(), PnlOverride.id.desc())
            .all())
    out = {s: None for s in VALID_SLOTS}
    for r in rows:
        # DESC order means the first hit per slot is the newest active one.
        if out.get(r.slot) is None:
            out[r.slot] = r
    return out


def store_override(values: dict, *, slot: str = "current",
                   source: str, filename: str | None,
                   source_path: str | None, file_sha256: str,
                   uploaded_by: str | None = None) -> PnlOverride:
    """Insert a new active override for the given slot; atomically deactivate any
    previously active rows in the same slot. History preserved as inactive rows."""
    _validate_slot(slot)
    PnlOverride.query.filter_by(slot=slot, is_active=True).update(
        {"is_active": False}, synchronize_session=False
    )
    row = PnlOverride(
        slot=slot,
        values=values,
        source=source,
        filename=filename,
        source_path=source_path,
        sheet_name=REQUIRED_SHEET,
        uploaded_by=uploaded_by or current_user_label(),
        file_sha256=file_sha256,
        is_active=True,
    )
    db.session.add(row)
    db.session.commit()
    return row


def deactivate_latest(slot: str = "current") -> bool:
    """Soft-clear the active override for a given slot. Returns True if one existed."""
    _validate_slot(slot)
    row = get_active_override(slot)
    if row is None:
        return False
    row.is_active = False
    db.session.commit()
    return True
