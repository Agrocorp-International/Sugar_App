"""
Shared Physical/Futures PNL and Exposure computation for the dashboard summary.
Loads all DB data once and reads the Excel file once for both Raws and Whites sheets.
"""
import openpyxl
from routes.raws import (
    _futures_pricing, _hedged, _purchase_status, _sales_status,
    _long_physical_pol, _short_physical_pol,
    _load_trade_maps, _load_settlement_prices, _load_futures_pnl_map,
    _load_whites_spread_maps, EXCEL_PATH,
)


def compute_all_pnl_totals(source='sett1'):
    """Return (raws_physical, whites_physical, raws_futures, whites_futures, raws_exposure, whites_exposure).

    Performs all DB queries once and opens the Excel file once.
    Any value is None if the data needed to compute it is missing.

    ``source`` selects sett-1 (default) or live prices/deltas, with silent
    fallback to sett-1 for any contract whose live value is missing.

    Exposure definitions:
      raws_exposure   = Σ (Purchase Exposure + Sales Exposure + Pol Exposure + Options Delta) per Raws row
      whites_exposure = Σ Overall Exposure (Purchase Exposure + Sales Exposure) per Whites row
    """
    # --- DB queries (done once) ---
    settlement_prices = _load_settlement_prices(source)
    (purchase_hedges, sales_hedges, purchase_price_total, sales_price_total,
     purchase_pol_hedges, sales_pol_hedges, options_delta, spread_price_total) = _load_trade_maps('Raws')
    (whites_purchase_hedges, whites_sales_hedges, whites_purchase_price_total, whites_sales_price_total,
     whites_purchase_pol_hedges, whites_sales_pol_hedges, _, _) = _load_trade_maps('Whites')
    whites_spreads_map, whites_spread_price_map, whites_futures_price_total = _load_whites_spread_maps(source)
    futures_pnl_map = _load_futures_pnl_map(settlement_prices, 'Raws')
    whites_futures_pnl_map = _load_futures_pnl_map(settlement_prices, 'Whites')

    raws_physical = None
    whites_physical = None
    raws_futures = None
    whites_futures = None
    raws_exposure = None
    whites_exposure = None

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        try:
            # --- Raws sheet ---
            ws = wb["Raws"]
            it = ws.iter_rows(values_only=True)
            next(it)  # group header
            raw_headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(it))]
            for row in it:
                if all(v is None for v in row):
                    continue
                d = dict(zip(raw_headers, row))
                if d.get("Shipment Period") is None:
                    continue

                agp = str(d.get("AGP") or "").strip()
                ags = str(d.get("AGS") or "").strip()

                # Futures PNL (same as table footer)
                raws_futures = (raws_futures or 0) + futures_pnl_map.get(agp, 0) + futures_pnl_map.get(ags, 0)

                # Physical PNL
                purchase_contract = str(d.get("Purchase Contract") or "")
                sales_contract = str(d.get("Sales Contract") or "")
                pc_norm = purchase_contract.replace(" ", "").upper()
                sc_norm = sales_contract.replace(" ", "").upper()

                qty_long = d.get("Qty Long") or 0
                qty_short = d.get("Qty Short") or 0
                long_physical_qty = int(float(qty_long) / 50.8 + 0.5) if qty_long else None
                short_physical_qty = -int(float(qty_short) / 50.8 + 0.5) if qty_short else None

                ph = purchase_hedges.get(agp)
                sh_h = sales_hedges.get(ags)
                purchase_hedges_val = round(ph) if ph is not None else None
                sales_hedges_val = round(sh_h) if sh_h is not None else None

                purchase_hedged = _hedged(purchase_hedges_val, long_physical_qty)
                sales_hedged = _hedged(sales_hedges_val, short_physical_qty)

                purchase_fp = _futures_pricing(
                    purchase_hedged,
                    settlement_prices.get(pc_norm),
                    purchase_hedges_val,
                    long_physical_qty,
                    purchase_price_total.get((pc_norm, agp)),
                )

                actual_pol = d.get("Actual Pol")
                if actual_pol is not None:
                    pol = float(actual_pol)
                elif (d.get("Status") or "").strip() == "Washout":
                    pol = 0.0375
                else:
                    pol = 0.042

                elevation = float(d.get("Elevation") or 0)
                purchase_terms = d.get("Purchase Terms")
                purchase_units = d.get("Purchase Units")
                purchase_input = d.get("Purchase Input")

                if purchase_terms == "Basis":
                    if purchase_units == "¢/lb":
                        base = ((purchase_fp + purchase_input) * 22.0462 * (1 + pol)) if (purchase_fp is not None and purchase_input is not None) else None
                    else:
                        base = ((purchase_fp * 22.0462) + purchase_input) if (purchase_fp is not None and purchase_input is not None) else None
                elif purchase_terms == "Flat":
                    base = purchase_input
                else:
                    base = None

                purchase_price = (base - elevation) if (base is not None and d.get("Purchase Incoterm") == "FCA") else base
                purchase_cost = (
                    (purchase_price or 0) +
                    float(d.get("Freight") or 0) +
                    float(d.get("Insurance") or 0) +
                    float(d.get("Financing") or 0) +
                    float(d.get("Misc") or 0)
                ) if purchase_price is not None else None

                sales_fp = _futures_pricing(
                    sales_hedged,
                    settlement_prices.get(sc_norm),
                    sales_hedges_val,
                    short_physical_qty,
                    sales_price_total.get((sc_norm, ags)),
                )
                sales_terms = d.get("Sales Terms")
                sales_units = d.get("Sales Units")
                sales_input = d.get("Sales Input")

                if sales_terms == "Basis":
                    if sales_units == "¢/lb":
                        sales_base = ((sales_fp + sales_input) * 22.0462 * (1 + pol)) if (sales_fp is not None and sales_input is not None) else None
                    else:
                        sales_base = ((sales_fp * 22.0462) + sales_input) if (sales_fp is not None and sales_input is not None) else None
                elif sales_terms == "Flat":
                    sales_base = sales_input
                else:
                    sales_base = None

                sales_price = (sales_base - elevation) if (sales_base is not None and d.get("Sales Incoterm") == "FCA") else sales_base

                if sales_price is not None and purchase_cost is not None:
                    qty = max(float(qty_long), float(qty_short))
                    raws_physical = (raws_physical or 0) + (sales_price - purchase_cost) * qty

                # --- Raws Exposure ---
                purchase_status = _purchase_status(purchase_terms, purchase_hedges_val, long_physical_qty)
                sales_status = _sales_status(ags, sales_terms, sales_hedges_val, short_physical_qty)

                if purchase_terms == "Basis":
                    pe = ((long_physical_qty or 0) + (purchase_hedges_val or 0)) if purchase_status == "Over Priced" else 0
                elif purchase_terms == "Flat":
                    pe = (long_physical_qty or 0) + (purchase_hedges_val or 0)
                else:
                    pe = 0

                if sales_terms == "Basis":
                    se = ((short_physical_qty or 0) + (sales_hedges_val or 0)) if sales_status == "Over Priced" else 0
                elif sales_terms == "Flat":
                    se = (short_physical_qty or 0) + (sales_hedges_val or 0)
                else:
                    se = 0

                long_pol = _long_physical_pol(purchase_units, purchase_status, long_physical_qty, pol, purchase_hedges_val) or 0
                short_pol = _short_physical_pol(sales_units, sales_status, short_physical_qty, pol, sales_hedges_val) or 0
                pol_exp = long_pol + short_pol + (purchase_pol_hedges.get(agp) or 0) + (sales_pol_hedges.get(ags) or 0)

                od = options_delta.get(agp, 0) + options_delta.get(ags, 0)
                od_val = round(od) if od else 0

                raws_exposure = (raws_exposure or 0) + pe + se + pol_exp + od_val

            # --- Whites sheet ---
            ws = wb["Whites"]
            it = ws.iter_rows(values_only=True)
            next(it)  # group header
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(it))]
            for row in it:
                if all(v is None for v in row):
                    continue
                d = dict(zip(headers, row))
                if d.get("Shipment Month") is None:
                    continue

                agp = str(d.get("AGP") or "").strip()
                ags = str(d.get("AGS") or "").strip()

                # Futures PNL (same as table footer)
                whites_futures = (whites_futures or 0) + whites_futures_pnl_map.get(agp, 0) + whites_futures_pnl_map.get(ags, 0)

                # Physical PNL
                pc = str(d.get("Purchase Contract") or "")
                sc = str(d.get("Sales Contract") or "")
                pc_norm = pc.replace(" ", "").upper()
                sc_norm = sc.replace(" ", "").upper()

                purchase_hedged_in_sb = "SB" in pc.upper()
                sales_hedged_in_sb = "SB" in sc.upper()

                qty_long = d.get("Qty Long")
                qty_short = d.get("Qty Short")
                divisor_p = 50.8 if purchase_hedged_in_sb else 50
                divisor_s = 50.8 if sales_hedged_in_sb else 50
                long_physical_qty = int(float(qty_long) / divisor_p + 0.5) if qty_long else None
                short_physical_qty = -int(float(qty_short) / divisor_s + 0.5) if qty_short else None

                ph = (whites_purchase_hedges.get(agp) or 0) + (whites_purchase_pol_hedges.get(agp) or 0)
                sh = (whites_sales_hedges.get(ags) or 0) + (whites_sales_pol_hedges.get(ags) or 0)
                purchase_hedges_val = round(ph) if agp else None
                sales_hedges_val = round(sh) if ags else None

                purchase_hedged = _hedged(purchase_hedges_val, long_physical_qty)
                sales_hedged = _hedged(sales_hedges_val, short_physical_qty)

                purchase_fp = _futures_pricing(
                    purchase_hedged,
                    settlement_prices.get(pc_norm) if pc else None,
                    purchase_hedges_val,
                    long_physical_qty,
                    whites_futures_price_total.get((pc_norm, agp)),
                )

                p_terms = d.get("Purchase Terms")
                s_terms = d.get("Sales Terms")
                purchase_input = d.get("Purchase Input")
                sales_input = d.get("Sales Input")

                if p_terms == "Basis":
                    factor = 22.0462 if purchase_hedged_in_sb else 1
                    purchase_price = (factor * purchase_fp + purchase_input) if (purchase_fp is not None and purchase_input is not None) else None
                elif p_terms == "Flat":
                    purchase_price = purchase_input
                else:
                    purchase_price = None

                cnf = (
                    purchase_price +
                    float(d.get("Freight") or 0) +
                    float(d.get("Insurance") or 0) +
                    float(d.get("Financing") or 0) +
                    float(d.get("Misc") or 0)
                ) if purchase_price is not None else None

                sales_fp = _futures_pricing(
                    sales_hedged,
                    settlement_prices.get(sc_norm) if sc else None,
                    sales_hedges_val,
                    short_physical_qty,
                    whites_futures_price_total.get((sc_norm, ags)),
                )

                if s_terms == "Basis":
                    factor = 22.0462 if sales_hedged_in_sb else 1
                    sales_price = (factor * sales_fp + sales_input) if (sales_fp is not None and sales_input is not None) else None
                elif s_terms == "Flat":
                    sales_price = sales_input
                else:
                    sales_price = None

                if sales_price is not None and cnf is not None:
                    qty = max(float(qty_long or 0), float(qty_short or 0))
                    whites_physical = (whites_physical or 0) + (sales_price - cnf) * qty

                # --- Whites Exposure (Overall Exposure = Purchase Exposure + Sales Exposure) ---
                purchase_status_w = _purchase_status(p_terms, purchase_hedges_val, long_physical_qty)
                sales_status_w = _sales_status(ags, s_terms, sales_hedges_val, short_physical_qty)

                if p_terms == "Basis":
                    pe_w = ((long_physical_qty or 0) + (purchase_hedges_val or 0)) if purchase_status_w == "Over Priced" else 0
                elif p_terms == "Flat":
                    pe_w = (long_physical_qty or 0) + (purchase_hedges_val or 0)
                else:
                    pe_w = 0

                if s_terms == "Basis":
                    se_w = ((short_physical_qty or 0) + (sales_hedges_val or 0)) if sales_status_w == "Over Priced" else 0
                elif s_terms == "Flat":
                    se_w = (short_physical_qty or 0) + (sales_hedges_val or 0)
                else:
                    se_w = 0

                whites_exposure = (whites_exposure or 0) + pe_w + se_w

        finally:
            wb.close()
    except Exception:
        pass

    return raws_physical, whites_physical, raws_futures, whites_futures, raws_exposure, whites_exposure
