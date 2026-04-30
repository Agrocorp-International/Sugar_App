import logging
import os
from flask import Blueprint, render_template, request, jsonify
import openpyxl
from pathlib import Path
from models.db import db, MarketPrice, TradePosition, PhysicalDeal
from routes.positions import build_contract_key, LOT_MULTIPLIERS as _MULTIPLIERS
from services.request_cache import get_all_positions, get_all_market_prices

logger = logging.getLogger(__name__)

raws_bp = Blueprint("raws", __name__)

EXCEL_PATH = Path(__file__).parent.parent / "sugarm2m.xlsm"

# When true (default), an empty book auto-seeds from sugarm2m.xlsm on the
# first page load — convenient for fresh dev DBs, slow + risky in prod.
# Set EXCEL_AUTOSEED=false on Azure to disable; the explicit /api/seed and
# /api/upload endpoints (the page's "Sync from Excel" / "Override from
# Excel" buttons) remain available either way.
EXCEL_AUTOSEED = os.environ.get("EXCEL_AUTOSEED", "true").lower() != "false"

WHITES_INPUT_COLS = [
    "Supplier", "AGP", "Year", "Shipment Month", "Buyer", "AGS",
    "Qty Long", "Qty Short", "Purchase Exposure", "Sales Exposure", "Overall Exposure",
    "Long Physical Qty", "Short Physical Qty", "Long Physical Exposure", "Short Physical Exposure", "Purchase Hedges", "Sales Hedges", "Spreads", "Spread Value",
    "Purchase Contract", "Purchase Settlement", "Purchase Incoterm", "Purchase Terms", "Purchase Input",
    "Purchase Futures Pricing", "Purchase Basis", "Purchase Price",
    "Sales Contract", "Sales Settlement", "Sales Incoterm", "Sales Terms", "Sales Input",
    "Sales Futures Pricing", "Sales Basis", "Sales Price",
    "Freight", "Insurance", "Financing", "Misc", "CNF",
    "Physical PNL", "Futures PNL", "Total PNL",
    "Commodity", "Origin", "Destination",
    "Purchase Status", "Sales Status",
    "Purchase Hedged in SB", "Sales Hedged in SB",
    "Purchase Hedged", "Sales Hedged",
    "Notes", "Shipment Status",
]

INPUT_COLS = [
    "Supplier", "AGP", "Buyer", "AGS", "Shipment Period",
    "Qty Long", "Qty Short",
    "Purchase Contract", "Purchase Incoterm", "Purchase Terms",
    "Purchase Units", "Purchase Input",
    "Sales Contract", "Sales Incoterm", "Sales Terms",
    "Sales Units", "Sales Input",
    "Freight", "Insurance", "Financing", "Misc", "Elevation",
    "Origin", "Destination", "Actual Pol", "Vessel",
    "Notes", "Shipment Status", "Status",
]

# Calculated columns inserted alongside input columns — headers shown in gold.
# ALL_COLS defines display order; values for calculated cols are computed in index().
CALCULATED_COLS = {"Purchase Settlement", "Long Physical Qty", "Short Physical Qty", "Purchase Hedges", "Sales Hedges", "Long Physical Pol", "Short Physical Pol", "Purchase Pol Hedge", "Sales Pol Hedge", "Options Delta", "Spreads", "Spread Value", "Overall Exposure", "Purchase Exposure", "Sales Exposure", "Pol Exposure", "Spread Exposure", "Purchase Status", "Sales Status", "Purchase Hedged", "Sales Hedged", "Pol", "Purchase Price", "Purchase Futures Pricing", "Purchase Basis", "Purchase Cost (CNF)", "Sales Settlement", "Sales Price", "Sales Futures Pricing", "Sales Basis", "Long Physical Exposure", "Short Physical Exposure", "Physical PNL", "Futures PNL", "Total PNL"}

NUMERIC_COLS = {
    "Qty Long", "Qty Short",
    "Long Physical Qty", "Short Physical Qty",
    "Purchase Hedges", "Sales Hedges", "Long Physical Pol", "Short Physical Pol", "Purchase Pol Hedge", "Sales Pol Hedge", "Options Delta", "Spreads",
    "Spread Value", "Overall Exposure", "Purchase Exposure", "Sales Exposure", "Pol Exposure", "Spread Exposure", "Purchase Input", "Purchase Price", "Purchase Cost (CNF)", "Physical PNL", "Futures PNL", "Total PNL", "Sales Input", "Sales Price",
    "Purchase Settlement", "Purchase Futures Pricing",
    "Sales Settlement", "Sales Futures Pricing",
    "Long Physical Exposure", "Short Physical Exposure",
    "Pol", "Freight", "Insurance", "Financing", "Misc", "Elevation",
}

ALL_COLS = [
    "Supplier", "AGP", "Buyer", "AGS", "Shipment Period",
    "Qty Long", "Qty Short", "Overall Exposure", "Purchase Exposure", "Sales Exposure", "Pol Exposure", "Spread Exposure", "Long Physical Qty", "Short Physical Qty", "Long Physical Exposure", "Short Physical Exposure", "Purchase Hedges", "Sales Hedges", "Long Physical Pol", "Short Physical Pol", "Purchase Pol Hedge", "Sales Pol Hedge", "Options Delta", "Spreads", "Spread Value",
    "Purchase Contract", "Purchase Settlement", "Purchase Incoterm", "Purchase Terms", "Purchase Units", "Purchase Input", "Purchase Futures Pricing", "Purchase Basis", "Purchase Price",
    "Sales Contract", "Sales Settlement", "Sales Incoterm", "Sales Terms", "Sales Units", "Sales Input", "Sales Futures Pricing", "Sales Basis", "Sales Price",
    "Freight", "Insurance", "Financing", "Misc", "Elevation", "Purchase Cost (CNF)", "Physical PNL", "Futures PNL", "Total PNL",
    "Origin", "Destination", "Pol", "Actual Pol", "Vessel", "Purchase Status", "Sales Status", "Purchase Hedged", "Sales Hedged",
    "Notes", "Shipment Status", "Status",
]


def _purchase_status(terms, hedges, long_physical_qty):
    """Derive Purchase Status from terms, hedge position, and physical qty."""
    h = abs(hedges or 0)
    q = abs(long_physical_qty or 0)
    if terms == "Basis":
        if h == 0:             return "Unpriced"
        if h < q:              return "Partially Priced"
        if h == q:             return "Fully Priced"
        return                        "Over Priced"
    if terms == "Flat":
        if h == 0:             return "Unhedged"
        if h < q:              return "Partially Hedged"
        if h == q:             return "Fully Hedged"
        return                        "Over Hedged"
    return None


def _hedged(hedges, physical_qty):
    """Compare rounded absolute values of hedge lots and physical qty.

    Returns 0 if hedges is zero, True if they match, False if they don't.
    Used for both Purchase Hedged and Sales Hedged.
    """
    h = abs(round(hedges or 0))
    if h == 0:
        return 0
    q = abs(round(physical_qty or 0))
    return h == q


def _futures_pricing(flag, settlement, hedges, physical_qty, sum_price_total):
    """Compute futures pricing based on hedge status. Used for both Purchase and Sales.

    flag == 0      → settlement price (unhedged)
    flag is True   → abs(SUM(Price Total)) / hedges  (weighted avg futures price, fully hedged)
    flag is False  → blended price: hedged portion at futures avg, remainder at settlement
    """
    if flag is True:
        sum_pt = sum_price_total or 0
        h = abs(hedges or 0)
        if h == 0:
            return None
        return abs(sum_pt) / h

    if flag is False:
        sum_pt = sum_price_total or 0
        h = abs(hedges or 0)
        q = abs(physical_qty or 0)
        if q == 0 or settlement is None:
            return None
        return (abs(sum_pt) + (q - h) * settlement) / q

    # flag == 0 (integer): unhedged → return settlement
    return settlement


def _physical_exposure(terms, status, physical_qty, hedges):
    """Derive net physical exposure after accounting for hedge pricing method.

    Used for both Long Physical Exposure (purchase) and Short Physical Exposure (sales).
    Basis: return full qty if over-priced, otherwise return negative hedges (unpriced portion).
    Flat:  always return full physical qty.
    """
    if terms == "Basis":
        if status == "Over Priced":
            return physical_qty
        return -(hedges or 0)
    if terms == "Flat":
        return physical_qty
    return None


def _sales_status(ags, terms, hedges, short_physical_qty):
    """Derive Sales Status. AGS='Delivered' takes priority over all other logic."""
    if (ags or "").strip() == "Delivered":
        return "Delivered"
    h = abs(hedges or 0)
    q = abs(short_physical_qty or 0)
    if terms == "Basis":
        if h == 0:  return "Unpriced"
        if h < q:   return "Partially Priced"
        if h == q:  return "Fully Priced"
        return             "Over Priced"
    if terms == "Flat":
        if h == 0:  return "Unhedged"
        if h < q:   return "Partially Hedged"
        if h == q:  return "Fully Hedged"
        return             "Over Hedged"
    return None


def _long_physical_pol(purchase_units, purchase_status, long_physical_qty, pol, purchase_hedges):
    """Compute Long Physical Pol (lot-adjusted polarisation exposure).

    Mirrors the Excel IFS formula:
      ¢/lb + Unpriced        → 0
      ¢/lb + Partially Priced → round(abs(Purchase Hedges) × Pol)
      ¢/lb + Fully/Over Priced → round(Long Physical Qty × Pol)
      non-¢/lb               → round(Long Physical Qty × Pol)
    """
    q = abs(long_physical_qty or 0)
    if purchase_units == "¢/lb":
        if purchase_status == "Unpriced":
            return 0
        if purchase_status == "Partially Priced":
            return round(abs(purchase_hedges or 0) * pol)
        if purchase_status in ("Fully Priced", "Over Priced"):
            return round(q * pol)
        return None
    return round(q * pol) if q else None


def _short_physical_pol(sales_units, sales_status, short_physical_qty, pol, sales_hedges):
    """Compute Short Physical Pol (lot-adjusted polarisation exposure, sales side).

    Mirrors _long_physical_pol but uses Sales columns and includes the Delivered case.
    Short Physical Qty is negative, so results are negative (matching the Excel formula
    which uses the raw column value without ABS):
      ¢/lb + Unpriced          → 0
      ¢/lb + Partially Priced  → -round(abs(Sales Hedges) × Pol)
      ¢/lb + Fully/Over Priced → round(Short Physical Qty × Pol)  (negative)
      ¢/lb + Delivered         → round(Short Physical Qty × Pol)  (negative)
      non-¢/lb                 → round(Short Physical Qty × Pol)  (negative)
    """
    q = (short_physical_qty or 0)
    if sales_units == "¢/lb":
        if sales_status == "Unpriced":
            return 0
        if sales_status == "Partially Priced":
            return round(-abs(sales_hedges or 0) * pol)
        if sales_status in ("Fully Priced", "Over Priced", "Delivered"):
            return round(q * pol)
        return None
    return round(q * pol) if q else None


def _load_trade_maps(book='Raws'):
    """Single DB pass returning dicts for positions matching the given book.

    Args:
        book: value of book_parsed to filter on (e.g. 'Raws' or 'Whites').

    Returns:
        purchase_hedges:      {contract_xl (AGP…) → net lots}  (non-Pol trades)
        sales_hedges:         {contract_xl (AGS…) → net lots}  (non-Pol trades)
        purchase_price_total: {(normalised_contract, contract_xl) → SUM(Price Total)}
        sales_price_total:    {(normalised_contract, contract_xl) → SUM(Price Total)}
            where Price Total = (Long__c + Short__c) × Price__c per position.
        purchase_pol_hedges:  {contract_xl (AGP…) → net lots}  (Pol trades only)
        sales_pol_hedges:     {contract_xl (AGS…) → net lots}  (Pol trades only)
        options_delta:        {contract_xl (AGP… or AGS…) → delta-adjusted position}  (Options trades)
    """
    market = {mp.contract.replace(' ', '').upper(): mp for mp in get_all_market_prices()}
    purchase_hedges = {}
    sales_hedges = {}
    purchase_price_total = {}
    sales_price_total = {}
    purchase_pol_hedges = {}
    sales_pol_hedges = {}
    options_delta = {}
    spread_price_total = {}
    for pos in get_all_positions():
        d = pos.data
        if (pos.book_parsed or '').strip() != book or not pos.contract_xl:
            continue
        long_ = float(d.get('Long__c') or 0)
        short_ = float(d.get('Short__c') or 0)
        net = long_ + short_
        cx = pos.contract_xl
        instrument = (pos.instrument or '').strip()
        if instrument == 'Options':
            key = build_contract_key(d)
            mp = market.get(key)
            delta = mp.delta if (mp and mp.delta is not None) else 0
            options_delta[cx] = options_delta.get(cx, 0) + delta * net
        elif instrument == 'Spread':
            price = float(d.get('Price__c') or 0)
            spread_price_total[cx] = spread_price_total.get(cx, 0) + net * price
        elif instrument == 'Futures':
            if d.get('Trade_Code__c') == 'Pol':
                if cx.startswith('AGP'):
                    purchase_pol_hedges[cx] = purchase_pol_hedges.get(cx, 0) + net
                elif cx.startswith('AGS'):
                    sales_pol_hedges[cx] = sales_pol_hedges.get(cx, 0) + net
            else:
                price = float(d.get('Price__c') or 0)
                contract_norm = (d.get('Contract__c') or '').replace(' ', '').upper()
                key = (contract_norm, cx)
                if cx.startswith('AGP'):
                    purchase_hedges[cx] = purchase_hedges.get(cx, 0) + net
                    purchase_price_total[key] = purchase_price_total.get(key, 0) + net * price
                elif cx.startswith('AGS'):
                    sales_hedges[cx] = sales_hedges.get(cx, 0) + net
                    sales_price_total[key] = sales_price_total.get(key, 0) + net * price
    return purchase_hedges, sales_hedges, purchase_price_total, sales_price_total, purchase_pol_hedges, sales_pol_hedges, options_delta, spread_price_total


def _load_settlement_prices(source='sett1'):
    """Return dict mapping normalised contract key (e.g. 'SBH26') → price from DB.

    ``source`` selects sett-1 (default) or live prices. When source='live',
    falls back to settlement for any contract whose live_price is None.
    Function name kept for backward compatibility with downstream importers.
    """
    from services.price_source import load_price_map
    pm, _ = load_price_map(source)
    return pm


def _load_spreads_map(book='Raws', source='sett1'):
    """Return {contract_xl → sum(spread_position)} for given book, Trade Code != Pol.

    Spread Position per trade mirrors the /positions page logic:
      - 0 if no spread contract
      - 0 if contract[-3:] == spread[-3:]  (same expiry, not a spread)
      - else (Long + Short) × delta

    ``source`` selects sett-1 (default) or live deltas with fallback.
    """
    from services.price_source import load_delta_map
    delta_by_key, _ = load_delta_map(source)
    result = {}
    for pos in get_all_positions():
        d = pos.data
        if (pos.book_parsed or '').strip() != book or d.get('Trade_Code__c') == 'Pol' or not pos.contract_xl:
            continue
        if (pos.instrument or '').strip() != 'Spread':
            continue
        spread = (pos.spread or '').strip()
        if not spread:
            spread_pos = 0.0
        else:
            contract = (d.get('Contract__c') or '').replace(' ', '')
            if contract[-3:] == spread[-3:]:
                spread_pos = 0.0
            else:
                key = build_contract_key(d)
                delta = delta_by_key.get(key)
                ls = float(d.get('Long__c') or 0) + float(d.get('Short__c') or 0)
                spread_pos = ls * delta if delta is not None else None
        if spread_pos is not None:
            result[pos.contract_xl] = result.get(pos.contract_xl, 0) + spread_pos
    return result


def _load_whites_spread_maps(source='sett1'):
    """Return (spreads_map, spread_price_map, futures_price_total) for Book=Whites.

    spreads_map:         {contract_xl → sum(spread_position)}           (all instruments)
    spread_price_map:    {contract_xl → sum(Price Total)}               (Spread instrument only)
    futures_price_total: {(contract_norm, contract_xl) → sum(Price Total)} (Futures instrument only)
    Spread Position logic mirrors /positions: 0 if no Strategy, 0 if same expiry, else (L+S)×delta.

    ``source`` selects sett-1 (default) or live deltas with fallback.
    """
    from services.price_source import load_delta_map
    delta_by_key, _ = load_delta_map(source)
    spreads_map = {}
    spread_price_map = {}
    futures_price_total = {}
    for pos in get_all_positions():
        d = pos.data
        if (pos.book_parsed or '').strip() != 'Whites' or not pos.contract_xl:
            continue
        cx = pos.contract_xl
        long_ = float(d.get('Long__c') or 0)
        short_ = float(d.get('Short__c') or 0)
        net = long_ + short_
        instrument = (pos.instrument or '').strip()
        if instrument == 'Spread':
            spread = (pos.spread or '').strip()
            if not spread:
                spread_pos = 0.0
            else:
                contract = (d.get('Contract__c') or '').replace(' ', '')
                if contract[-3:] == spread[-3:]:
                    spread_pos = 0.0
                else:
                    key = build_contract_key(d)
                    delta = delta_by_key.get(key)
                    spread_pos = net * delta if delta is not None else None
            if spread_pos is not None:
                spreads_map[cx] = spreads_map.get(cx, 0) + spread_pos
        price = float(d.get('Price__c') or 0)
        if instrument == 'Spread':
            spread_price_map[cx] = spread_price_map.get(cx, 0) + net * price
        elif instrument == 'Futures':
            contract_norm = (d.get('Contract__c') or '').replace(' ', '').upper()
            fkey = (contract_norm, cx)
            futures_price_total[fkey] = futures_price_total.get(fkey, 0) + net * price
    return spreads_map, spread_price_map, futures_price_total



def _load_futures_pnl_map(settlement_prices, book='Raws'):
    """Return {contract_xl → sum(net_pnl)} for positions matching the given book.

    Net PNL = (Settlement - Price) × Lots × Multiplier + Broker Commission.
    Used to compute Futures PNL on the Raws page by summing AGP and AGS contract_xl buckets.
    """
    result = {}
    for pos in get_all_positions():
        d = pos.data
        if (pos.book_parsed or '').strip() != book or not pos.contract_xl:
            continue
        key = build_contract_key(d)
        settlement = settlement_prices.get(key)
        is_option = bool(d.get("Put_Call_2__c") and d.get("Strike__c") is not None)
        if settlement is None and is_option:
            settlement = 0
        price = d.get("Price__c")
        lots = float(d.get("Long__c") or 0) + float(d.get("Short__c") or 0)
        multiplier = _MULTIPLIERS.get(d.get("Commodity_Name__c") or "", 0)
        pnl = (settlement - float(price)) * lots * multiplier if (settlement is not None and price is not None) else 0
        commission = pos.commission
        result[pos.contract_xl] = result.get(pos.contract_xl, 0) + pnl + commission
    return result



def _parse_workbook(wb):
    """Parse a workbook into PhysicalDeal row payloads for both books.

    Returns (raws_payloads, whites_payloads, warnings) where each payload is
    a list of dicts ready for bulk_insert_mappings.
    Raises ValueError on missing sheets or headers.
    """
    sheet_names = wb.sheetnames
    missing = [s for s in ("Raws", "Whites") if s not in sheet_names]
    if missing:
        raise ValueError(f"Missing required sheet(s): {', '.join(missing)}")

    warnings = []

    # --- Raws ---
    ws = wb["Raws"]
    it = ws.iter_rows(values_only=True)
    next(it)  # skip group header row
    raw_headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(it))]
    missing_raws = [c for c in INPUT_COLS if c not in raw_headers]
    if missing_raws:
        raise ValueError(f"Raws sheet missing headers: {missing_raws}")

    raws_payloads = []
    idx = 0
    for row in it:
        if all(v is None for v in row):
            continue
        d = dict(zip(raw_headers, row))
        if d.get("Shipment Period") is None:
            continue
        data = {col: d.get(col) for col in INPUT_COLS}
        if "Physical PNL" in raw_headers and d.get("Physical PNL") is not None:
            data["_excel_physical_pnl"] = d["Physical PNL"]
        raws_payloads.append({"book": "Raws", "row_index": idx, "data": data})
        idx += 1

    # --- Whites ---
    ws_w = wb["Whites"]
    it_w = ws_w.iter_rows(values_only=True)
    next(it_w)  # skip group header row
    whites_headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(it_w))]
    # Whites has many calculated cols in WHITES_INPUT_COLS that may not be in Excel headers.
    # Only validate the core input columns that must exist in the sheet.
    WHITES_REQUIRED_HEADERS = {"Supplier", "AGP", "Buyer", "AGS", "Shipment Month", "Qty Long", "Qty Short",
                               "Purchase Contract", "Sales Contract", "Purchase Terms", "Sales Terms",
                               "Purchase Input", "Sales Input"}
    missing_whites = [c for c in WHITES_REQUIRED_HEADERS if c not in whites_headers]
    if missing_whites:
        raise ValueError(f"Whites sheet missing headers: {missing_whites}")

    whites_payloads = []
    idx = 0
    for row in it_w:
        if all(v is None for v in row):
            continue
        d = dict(zip(whites_headers, row))
        if d.get("Shipment Month") is None:
            continue
        data = {col: d.get(col) for col in WHITES_INPUT_COLS if col in d}
        whites_payloads.append({"book": "Whites", "row_index": idx, "data": data})
        idx += 1

    return raws_payloads, whites_payloads, warnings


def _store_payloads(raws_payloads, whites_payloads):
    """Delete existing PhysicalDeal rows for both books and insert new ones.
    Single transaction — parse first, replace second, commit once.
    """
    PhysicalDeal.query.filter(PhysicalDeal.book.in_(["Raws", "Whites"])).delete(synchronize_session=False)
    if raws_payloads:
        db.session.bulk_insert_mappings(PhysicalDeal, raws_payloads)
    if whites_payloads:
        db.session.bulk_insert_mappings(PhysicalDeal, whites_payloads)
    db.session.commit()


def _seed_book_from_excel(book):
    """Auto-seed a single book from EXCEL_PATH if the file exists.
    Used as fallback when DB is empty for a book.
    """
    if not EXCEL_PATH.exists():
        return
    logger.info("Auto-seed fallback triggered for book=%s from %s", book, EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        raws_payloads, whites_payloads, _ = _parse_workbook(wb)
    finally:
        wb.close()
    if book == "Raws" and raws_payloads:
        PhysicalDeal.query.filter_by(book="Raws").delete(synchronize_session=False)
        db.session.bulk_insert_mappings(PhysicalDeal, raws_payloads)
        db.session.commit()
        logger.info("Auto-seeded %d Raws rows", len(raws_payloads))
    elif book == "Whites" and whites_payloads:
        PhysicalDeal.query.filter_by(book="Whites").delete(synchronize_session=False)
        db.session.bulk_insert_mappings(PhysicalDeal, whites_payloads)
        db.session.commit()
        logger.info("Auto-seeded %d Whites rows", len(whites_payloads))


@raws_bp.route("/raws/api/seed", methods=["POST"])
def api_seed():
    """Seed DB from the on-disk sugarm2m.xlsm file. Replaces both books."""
    if not EXCEL_PATH.exists():
        return jsonify({"error": f"File not found: {EXCEL_PATH}"}), 400
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        try:
            raws_payloads, whites_payloads, warnings = _parse_workbook(wb)
        finally:
            wb.close()
        _store_payloads(raws_payloads, whites_payloads)
        logger.info("Seed complete: %d Raws, %d Whites rows", len(raws_payloads), len(whites_payloads))
        return jsonify({"ok": True, "raws": len(raws_payloads), "whites": len(whites_payloads), "warnings": warnings})
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Seed failed")
        return jsonify({"error": str(e)}), 500


@raws_bp.route("/raws/api/upload", methods=["POST"])
def api_upload():
    """Upload an Excel file to replace both books. Emergency override."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            raws_payloads, whites_payloads, warnings = _parse_workbook(wb)
        finally:
            wb.close()
        _store_payloads(raws_payloads, whites_payloads)
        logger.info("Upload override complete: %d Raws, %d Whites rows", len(raws_payloads), len(whites_payloads))
        return jsonify({"ok": True, "raws": len(raws_payloads), "whites": len(whites_payloads), "warnings": warnings})
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Upload override failed")
        return jsonify({"error": str(e)}), 500


@raws_bp.route("/raws")
def index():
    from services.price_source import get_price_source
    price_source = get_price_source()
    rows = []
    whites_rows = []
    error = None
    try:
        settlement_prices = _load_settlement_prices(price_source)
        purchase_hedges, sales_hedges, purchase_price_total, sales_price_total, purchase_pol_hedges, sales_pol_hedges, options_delta, spread_price_total = _load_trade_maps('Raws')
        futures_pnl_map = _load_futures_pnl_map(settlement_prices, 'Raws')
        spreads_map = _load_spreads_map('Raws', price_source)
        whites_purchase_hedges, whites_sales_hedges, whites_purchase_price_total, whites_sales_price_total, whites_purchase_pol_hedges, whites_sales_pol_hedges, whites_options_delta, whites_spread_price_total_map = _load_trade_maps('Whites')
        whites_futures_pnl_map = _load_futures_pnl_map(settlement_prices, 'Whites')
        whites_spreads_map, whites_spread_price_map, whites_futures_price_total = _load_whites_spread_maps(price_source)

        # Read Raws deals from DB (auto-seed from Excel if empty + flag on).
        raws_deals = PhysicalDeal.query.filter_by(book="Raws").order_by(PhysicalDeal.row_index).all()
        if not raws_deals and EXCEL_AUTOSEED:
            _seed_book_from_excel("Raws")
            raws_deals = PhysicalDeal.query.filter_by(book="Raws").order_by(PhysicalDeal.row_index).all()

        for deal in raws_deals:
            d = deal.data
            row_data = {col: d.get(col) for col in INPUT_COLS}
            row_data["_deal_id"] = deal.id
            row_data["_highlight"] = d.get("_highlight")
            # Inject calculated columns
            purchase_contract_norm = str(d.get("Purchase Contract") or "").replace(' ', '').upper()
            sales_contract_norm = str(d.get("Sales Contract") or "").replace(' ', '').upper()
            row_data["Purchase Settlement"] = settlement_prices.get(purchase_contract_norm)
            row_data["Sales Settlement"] = settlement_prices.get(sales_contract_norm)
            qty_long = d.get("Qty Long") or 0
            qty_short = d.get("Qty Short") or 0
            row_data["Long Physical Qty"] = int(float(qty_long) / 50.8 + 0.5) if qty_long else None
            row_data["Short Physical Qty"] = -int(float(qty_short) / 50.8 + 0.5) if qty_short else None
            agp = str(d.get("AGP") or "").strip()
            ags = str(d.get("AGS") or "").strip()
            ph = purchase_hedges.get(agp)
            sh = sales_hedges.get(ags)
            row_data["Purchase Hedges"] = round(ph) if ph is not None else None
            row_data["Sales Hedges"] = round(sh) if sh is not None else None
            row_data["Purchase Hedged"] = _hedged(row_data["Purchase Hedges"], row_data["Long Physical Qty"])
            row_data["Sales Hedged"] = _hedged(row_data["Sales Hedges"], row_data["Short Physical Qty"])
            pt_key = (purchase_contract_norm, agp)
            row_data["Purchase Futures Pricing"] = _futures_pricing(
                row_data["Purchase Hedged"],
                row_data["Purchase Settlement"],
                row_data["Purchase Hedges"],
                row_data["Long Physical Qty"],
                purchase_price_total.get(pt_key),
            )
            st_key = (sales_contract_norm, ags)
            purchase_contract = d.get("Purchase Contract")
            purchase_input = d.get("Purchase Input")
            purchase_fp = row_data["Purchase Futures Pricing"]
            if not purchase_contract:
                row_data["Purchase Basis"] = "-"
            elif d.get("Purchase Terms") == "Basis":
                row_data["Purchase Basis"] = purchase_input
            elif d.get("Purchase Terms") == "Flat":
                row_data["Purchase Basis"] = (purchase_input - purchase_fp) if (purchase_input is not None and purchase_fp is not None) else None
            else:
                row_data["Purchase Basis"] = None
            actual_pol = d.get("Actual Pol")
            if actual_pol is not None and actual_pol != "":
                pol = float(actual_pol)
            elif (d.get("Status") or "").strip() == "Washout":
                pol = 0.0375
            else:
                pol = 0.042
            row_data["Pol"] = pol
            elevation = float(d.get("Elevation") or 0)
            purchase_terms = d.get("Purchase Terms")
            purchase_units = d.get("Purchase Units")
            if purchase_terms == "Basis":
                if purchase_units == "¢/lb":
                    base = ((purchase_fp + purchase_input) * 22.0462 * (1 + pol)) if (purchase_fp is not None and purchase_input is not None) else None
                else:
                    base = ((purchase_fp * 22.0462) + purchase_input) if (purchase_fp is not None and purchase_input is not None) else None
            elif purchase_terms == "Flat":
                base = purchase_input
            else:
                base = None
            if base is not None and d.get("Purchase Incoterm") == "FCA":
                row_data["Purchase Price"] = base - elevation
            else:
                row_data["Purchase Price"] = base
            pp = row_data["Purchase Price"]
            row_data["Purchase Cost (CNF)"] = (
                (pp or 0) +
                float(d.get("Freight") or 0) +
                float(d.get("Insurance") or 0) +
                float(d.get("Financing") or 0) +
                float(d.get("Misc") or 0)
            ) if pp is not None else None
            row_data["Sales Futures Pricing"] = _futures_pricing(
                row_data["Sales Hedged"],
                row_data["Sales Settlement"],
                row_data["Sales Hedges"],
                row_data["Short Physical Qty"],
                sales_price_total.get(st_key),
            )
            sales_contract = d.get("Sales Contract")
            sales_input = d.get("Sales Input")
            sales_fp = row_data["Sales Futures Pricing"]
            if not sales_contract:
                row_data["Sales Basis"] = "-"
            elif d.get("Sales Terms") == "Basis":
                row_data["Sales Basis"] = sales_input
            elif d.get("Sales Terms") == "Flat":
                row_data["Sales Basis"] = (sales_input - sales_fp) if (sales_input is not None and sales_fp is not None) else None
            else:
                row_data["Sales Basis"] = None
            sales_terms = d.get("Sales Terms")
            sales_units = d.get("Sales Units")
            if sales_terms == "Basis":
                if sales_units == "¢/lb":
                    sales_base = ((sales_fp + sales_input) * 22.0462 * (1 + pol)) if (sales_fp is not None and sales_input is not None) else None
                else:
                    sales_base = ((sales_fp * 22.0462) + sales_input) if (sales_fp is not None and sales_input is not None) else None
            elif sales_terms == "Flat":
                sales_base = sales_input
            else:
                sales_base = None
            if sales_base is not None and d.get("Sales Incoterm") == "FCA":
                row_data["Sales Price"] = sales_base - elevation
            else:
                row_data["Sales Price"] = sales_base
            sp = row_data["Sales Price"]
            pc = row_data["Purchase Cost (CNF)"]
            if sp is not None and pc is not None:
                qty = max(float(d.get("Qty Long") or 0), float(d.get("Qty Short") or 0))
                row_data["Physical PNL"] = (sp - pc) * qty
            else:
                row_data["Physical PNL"] = None
            xl_pnl = d.get("_excel_physical_pnl")
            computed_pnl = row_data["Physical PNL"]
            if xl_pnl is not None and computed_pnl is not None:
                try:
                    if abs(float(computed_pnl) - float(xl_pnl)) > 0.01:
                        row_data["_pnl_mismatch"] = True
                        row_data["_excel_physical_pnl"] = xl_pnl
                except (TypeError, ValueError):
                    pass
            row_data["Futures PNL"] = (
                futures_pnl_map.get(agp, 0) + futures_pnl_map.get(ags, 0)
            )
            physical = row_data["Physical PNL"]
            row_data["Total PNL"] = (physical + row_data["Futures PNL"]) if physical is not None else None
            row_data["Purchase Status"] = _purchase_status(
                d.get("Purchase Terms"),
                row_data["Purchase Hedges"],
                row_data["Long Physical Qty"],
            )
            row_data["Long Physical Pol"] = _long_physical_pol(
                d.get("Purchase Units"),
                row_data["Purchase Status"],
                row_data["Long Physical Qty"],
                pol,
                row_data["Purchase Hedges"],
            )
            row_data["Long Physical Exposure"] = _physical_exposure(
                d.get("Purchase Terms"),
                row_data["Purchase Status"],
                row_data["Long Physical Qty"],
                row_data["Purchase Hedges"],
            )
            purchase_terms = d.get("Purchase Terms")
            lpq = row_data["Long Physical Qty"] or 0
            ph_lots = row_data["Purchase Hedges"] or 0
            if purchase_terms == "Basis":
                row_data["Purchase Exposure"] = (lpq + ph_lots) if row_data["Purchase Status"] == "Over Priced" else 0
            elif purchase_terms == "Flat":
                row_data["Purchase Exposure"] = lpq + ph_lots
            else:
                row_data["Purchase Exposure"] = None
            row_data["Sales Status"] = _sales_status(
                d.get("AGS"),
                d.get("Sales Terms"),
                row_data["Sales Hedges"],
                row_data["Short Physical Qty"],
            )
            row_data["Short Physical Exposure"] = _physical_exposure(
                d.get("Sales Terms"),
                row_data["Sales Status"],
                row_data["Short Physical Qty"],
                row_data["Sales Hedges"],
            )
            sales_terms = d.get("Sales Terms")
            spq = row_data["Short Physical Qty"] or 0
            sh_lots = row_data["Sales Hedges"] or 0
            if sales_terms == "Basis":
                row_data["Sales Exposure"] = (spq + sh_lots) if row_data["Sales Status"] == "Over Priced" else 0
            elif sales_terms == "Flat":
                row_data["Sales Exposure"] = spq + sh_lots
            else:
                row_data["Sales Exposure"] = None
            row_data["Short Physical Pol"] = _short_physical_pol(
                d.get("Sales Units"),
                row_data["Sales Status"],
                row_data["Short Physical Qty"],
                pol,
                row_data["Sales Hedges"],
            )
            ph = purchase_pol_hedges.get(agp)
            row_data["Purchase Pol Hedge"] = round(ph) if ph is not None else None
            sh = sales_pol_hedges.get(ags)
            row_data["Sales Pol Hedge"] = round(sh) if sh is not None else None
            row_data["Pol Exposure"] = (
                (row_data["Long Physical Pol"] or 0) +
                (row_data["Short Physical Pol"] or 0) +
                (row_data["Purchase Pol Hedge"] or 0) +
                (row_data["Sales Pol Hedge"] or 0)
            )
            od = options_delta.get(agp, 0) + options_delta.get(ags, 0)
            row_data["Options Delta"] = round(od) if od else None
            sp_val = spreads_map.get(agp, 0) + spreads_map.get(ags, 0)
            row_data["Spreads"] = round(sp_val, 1) if sp_val else None
            spreads = row_data["Spreads"]
            if not spreads:
                row_data["Spread Value"] = "-"
            elif spreads > 0:
                pt = spread_price_total.get(agp, 0)
                row_data["Spread Value"] = pt / spreads
            else:
                pt = spread_price_total.get(ags, 0)
                row_data["Spread Value"] = pt / spreads
            base = 0 if d.get("Purchase Contract") == d.get("Sales Contract") else (row_data["Long Physical Qty"] or 0)
            row_data["Spread Exposure"] = base - abs(row_data["Spreads"] or 0)
            row_data["Overall Exposure"] = (
                (row_data["Purchase Exposure"] or 0) +
                (row_data["Sales Exposure"] or 0) +
                (row_data["Pol Exposure"] or 0) +
                (row_data["Options Delta"] or 0)
            )
            rows.append(row_data)

        # Read Whites deals from DB (auto-seed from Excel if empty + flag on).
        whites_deals = PhysicalDeal.query.filter_by(book="Whites").order_by(PhysicalDeal.row_index).all()
        if not whites_deals and EXCEL_AUTOSEED:
            _seed_book_from_excel("Whites")
            whites_deals = PhysicalDeal.query.filter_by(book="Whites").order_by(PhysicalDeal.row_index).all()

        for deal_w in whites_deals:
            d = deal_w.data
            row_w = {col: d.get(col) for col in WHITES_INPUT_COLS if col in d}
            row_w["_deal_id"] = deal_w.id
            row_w["_highlight"] = d.get("_highlight")
            pc = str(d.get("Purchase Contract") or "")
            sc = str(d.get("Sales Contract") or "")
            row_w["Purchase Settlement"] = settlement_prices.get(pc.replace(' ', '').upper()) if pc else None
            row_w["Sales Settlement"] = settlement_prices.get(sc.replace(' ', '').upper()) if sc else None
            row_w["Purchase Hedged in SB"] = "SB" in pc.upper()
            row_w["Sales Hedged in SB"] = "SB" in sc.upper()
            qty_long = d.get("Qty Long")
            qty_short = d.get("Qty Short")
            divisor_p = 50.8 if row_w["Purchase Hedged in SB"] else 50
            divisor_s = 50.8 if row_w["Sales Hedged in SB"] else 50
            row_w["Long Physical Qty"] = int(float(qty_long) / divisor_p + 0.5) if qty_long else None
            row_w["Short Physical Qty"] = -int(float(qty_short) / divisor_s + 0.5) if qty_short else None
            agp = str(d.get("AGP") or "").strip()
            ags = str(d.get("AGS") or "").strip()
            ph = (whites_purchase_hedges.get(agp) or 0) + (whites_purchase_pol_hedges.get(agp) or 0)
            sh = (whites_sales_hedges.get(ags) or 0) + (whites_sales_pol_hedges.get(ags) or 0)
            row_w["Purchase Hedges"] = round(ph) if agp else None
            row_w["Sales Hedges"] = round(sh) if ags else None
            row_w["Purchase Status"] = _purchase_status(
                d.get("Purchase Terms"), row_w["Purchase Hedges"], row_w["Long Physical Qty"]
            )
            row_w["Sales Status"] = _sales_status(
                ags, d.get("Sales Terms"), row_w["Sales Hedges"], row_w["Short Physical Qty"]
            )
            row_w["Purchase Hedged"] = _hedged(row_w["Purchase Hedges"], row_w["Long Physical Qty"])
            row_w["Sales Hedged"] = _hedged(row_w["Sales Hedges"], row_w["Short Physical Qty"])
            p_terms = d.get("Purchase Terms")
            s_terms = d.get("Sales Terms")
            lpq = row_w["Long Physical Qty"] or 0
            spq = row_w["Short Physical Qty"] or 0
            ph = row_w["Purchase Hedges"] or 0
            sh = row_w["Sales Hedges"] or 0
            if p_terms == "Basis":
                pe = (lpq + ph) if row_w["Purchase Status"] == "Over Priced" else 0
            elif p_terms == "Flat":
                pe = lpq + ph
            else:
                pe = None
            if s_terms == "Basis":
                se = (spq + sh) if row_w["Sales Status"] == "Over Priced" else 0
            elif s_terms == "Flat":
                se = spq + sh
            else:
                se = None
            row_w["Purchase Exposure"] = pe
            row_w["Sales Exposure"] = se
            row_w["Overall Exposure"] = (pe or 0) + (se or 0) if (pe is not None or se is not None) else None
            row_w["Long Physical Exposure"] = _physical_exposure(
                d.get("Purchase Terms"), row_w["Purchase Status"], row_w["Long Physical Qty"], row_w["Purchase Hedges"]
            )
            row_w["Short Physical Exposure"] = _physical_exposure(
                d.get("Sales Terms"), row_w["Sales Status"], row_w["Short Physical Qty"], row_w["Sales Hedges"]
            )
            w_spreads = (whites_spreads_map.get(agp) or 0) + (whites_spreads_map.get(ags) or 0)
            row_w["Spreads"] = round(w_spreads)
            if w_spreads == 0:
                row_w["Spread Value"] = "-"
            else:
                sp_price = whites_spread_price_map.get(agp) or 0
                row_w["Spread Value"] = sp_price / w_spreads
            pc_norm = pc.replace(' ', '').upper()
            pt_key = (pc_norm, agp)
            row_w["Purchase Futures Pricing"] = _futures_pricing(
                row_w["Purchase Hedged"],
                row_w["Purchase Settlement"],
                row_w["Purchase Hedges"],
                row_w["Long Physical Qty"],
                whites_futures_price_total.get(pt_key),
            )
            purchase_input = d.get("Purchase Input")
            purchase_fp = row_w["Purchase Futures Pricing"]
            if not pc:
                row_w["Purchase Basis"] = "-"
            elif p_terms == "Basis":
                row_w["Purchase Basis"] = purchase_input
            elif p_terms == "Flat":
                row_w["Purchase Basis"] = (purchase_input - purchase_fp) if (purchase_input is not None and purchase_fp is not None) else None
            else:
                row_w["Purchase Basis"] = None
            if p_terms == "Basis":
                factor = 22.0462 if row_w["Purchase Hedged in SB"] else 1
                row_w["Purchase Price"] = (factor * purchase_fp + purchase_input) if (purchase_fp is not None and purchase_input is not None) else None
            elif p_terms == "Flat":
                row_w["Purchase Price"] = purchase_input
            else:
                row_w["Purchase Price"] = None
            sc_norm = sc.replace(' ', '').upper()
            st_key = (sc_norm, ags)
            row_w["Sales Futures Pricing"] = _futures_pricing(
                row_w["Sales Hedged"],
                row_w["Sales Settlement"],
                row_w["Sales Hedges"],
                row_w["Short Physical Qty"],
                whites_futures_price_total.get(st_key),
            )
            sales_input = d.get("Sales Input")
            sales_fp = row_w["Sales Futures Pricing"]
            if s_terms == "Basis":
                row_w["Sales Basis"] = sales_input
            elif s_terms == "Flat":
                row_w["Sales Basis"] = (sales_input - sales_fp) if (sales_input is not None and sales_fp is not None) else None
            else:
                row_w["Sales Basis"] = None
            if s_terms == "Basis":
                factor = 22.0462 if row_w["Sales Hedged in SB"] else 1
                row_w["Sales Price"] = (factor * sales_fp + sales_input) if (sales_fp is not None and sales_input is not None) else None
            elif s_terms == "Flat":
                row_w["Sales Price"] = sales_input
            else:
                row_w["Sales Price"] = None
            pp = row_w.get("Purchase Price")
            freight = float(d.get("Freight") or 0)
            insurance = float(d.get("Insurance") or 0)
            financing = float(d.get("Financing") or 0)
            misc = float(d.get("Misc") or 0)
            row_w["CNF"] = (pp + freight + insurance + financing + misc) if pp is not None else None
            cnf = row_w["CNF"]
            sp_w = row_w.get("Sales Price")
            if sp_w is not None and cnf is not None:
                qty = max(float(qty_long or 0), float(qty_short or 0))
                row_w["Physical PNL"] = (sp_w - cnf) * qty
            else:
                row_w["Physical PNL"] = 0
            xl_pnl_w = d.get("Physical PNL")
            computed_pnl_w = row_w["Physical PNL"]
            if xl_pnl_w is not None and computed_pnl_w is not None:
                try:
                    if abs(float(computed_pnl_w) - float(xl_pnl_w)) > 0.01:
                        row_w["_pnl_mismatch"] = True
                        row_w["_excel_physical_pnl"] = xl_pnl_w
                except (TypeError, ValueError):
                    pass
            row_w["Futures PNL"] = (
                whites_futures_pnl_map.get(agp, 0) + whites_futures_pnl_map.get(ags, 0)
            )
            physical_w = row_w["Physical PNL"]
            row_w["Total PNL"] = physical_w + row_w["Futures PNL"]
            whites_rows.append(row_w)

    except Exception as e:
        error = str(e)
    PNL_COLS = {"Physical PNL", "Futures PNL", "Total PNL"}
    TOTAL_COLS = PNL_COLS | {
        "Qty Long", "Qty Short",
        "Overall Exposure", "Purchase Exposure", "Sales Exposure", "Pol Exposure", "Spread Exposure",
        "Long Physical Qty", "Short Physical Qty",
        "Long Physical Exposure", "Short Physical Exposure",
        "Purchase Hedges", "Sales Hedges",
        "Long Physical Pol", "Short Physical Pol",
        "Purchase Pol Hedge", "Sales Pol Hedge",
        "Options Delta", "Spreads",
    }
    totals = {col: sum(r[col] for r in rows if r.get(col) is not None) for col in TOTAL_COLS}
    WHITES_PNL_COLS = {"Physical PNL", "Futures PNL", "Total PNL"}
    WHITES_TOTAL_COLS = WHITES_PNL_COLS | {"Qty Long", "Qty Short", "Purchase Exposure", "Sales Exposure", "Overall Exposure", "Long Physical Qty", "Short Physical Qty", "Long Physical Exposure", "Short Physical Exposure", "Purchase Hedges", "Sales Hedges", "Spreads"}
    whites_totals = {col: sum(r[col] for r in whites_rows if isinstance(r.get(col), (int, float))) for col in WHITES_TOTAL_COLS}
    whites_calculated_cols = {"Purchase Hedged in SB", "Sales Hedged in SB", "Purchase Hedged", "Sales Hedged", "Long Physical Qty", "Short Physical Qty", "Long Physical Exposure", "Short Physical Exposure", "Purchase Hedges", "Sales Hedges", "Purchase Status", "Sales Status", "Purchase Exposure", "Sales Exposure", "Overall Exposure", "Spreads", "Spread Value", "Purchase Settlement", "Sales Settlement", "Purchase Futures Pricing", "Purchase Basis", "Purchase Price", "Sales Futures Pricing", "Sales Basis", "Sales Price", "CNF", "Physical PNL", "Futures PNL", "Total PNL"}
    whites_numeric_cols = {
        "Qty Long", "Qty Short",
        "Purchase Exposure", "Sales Exposure", "Overall Exposure",
        "Long Physical Qty", "Short Physical Qty",
        "Long Physical Exposure", "Short Physical Exposure",
        "Purchase Hedges", "Sales Hedges",
        "Spreads",
        "Purchase Settlement", "Sales Settlement",
        "Purchase Futures Pricing", "Purchase Basis", "Purchase Price",
        "Sales Futures Pricing", "Sales Basis", "Sales Price",
        "Purchase Input", "Sales Input",
        "Freight", "Insurance", "Financing", "Misc", "CNF",
        "Physical PNL", "Futures PNL", "Total PNL",
    }
    return render_template("raws.html", rows=rows, columns=ALL_COLS, calculated_cols=CALCULATED_COLS, numeric_cols=NUMERIC_COLS, pnl_cols=PNL_COLS, total_cols=TOTAL_COLS, totals=totals, whites_rows=whites_rows, whites_columns=WHITES_INPUT_COLS, whites_calculated_cols=whites_calculated_cols, whites_numeric_cols=whites_numeric_cols, whites_pnl_cols=WHITES_PNL_COLS, whites_total_cols=WHITES_TOTAL_COLS, whites_totals=whites_totals, error=error, price_source=price_source)


@raws_bp.route("/raws/api/update", methods=["POST"])
def api_update():
    """Update editable fields on physical deals. Validates field against deal's book."""
    from sqlalchemy.orm.attributes import flag_modified
    data = request.get_json()
    if not data:
        return jsonify({"error": "No JSON body"}), 400
    changes = data.get("changes", [])
    if not changes:
        return jsonify({"ok": True, "updated": 0})

    RAWS_ALLOWED = set(INPUT_COLS)
    WHITES_ALLOWED = set(WHITES_INPUT_COLS)

    _MONTHS = {"Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"}
    ALLOWED_VALUES = {
        "Shipment Period": _MONTHS,
        "Shipment Month":  _MONTHS,
        "Purchase Terms":  {"Basis","Flat"},
        "Sales Terms":     {"Basis","Flat"},
        "Purchase Units":  {"$/MT","c/lb"},
        "Sales Units":     {"$/MT","c/lb"},
    }
    UPPERCASE_FIELDS = {"Purchase Contract", "Sales Contract"}
    CONTRACT_FIELDS = {"Purchase Contract", "Sales Contract"}
    CONTRACT_MONTHS = set("FGHJKMNQUVXZ")
    import re as _re
    _contract_re = _re.compile(r"^(SB|SW)\s([FGHJKMNQUVXZ])(\d{2})$")

    try:
        for change in changes:
            record_id = change.get("record_id")
            field = change.get("field", "").strip()
            value = change.get("value")

            deal = db.session.get(PhysicalDeal, record_id)
            if not deal:
                continue

            allowed = RAWS_ALLOWED if deal.book == "Raws" else WHITES_ALLOWED
            if field not in allowed:
                return jsonify({"error": f"Disallowed field '{field}' for book '{deal.book}'"}), 400

            # Uppercase contract codes
            if field in UPPERCASE_FIELDS and isinstance(value, str):
                value = value.strip().upper()

            # Normalize & validate contract format: "SB K26"
            if field in CONTRACT_FIELDS and isinstance(value, str) and value:
                compact = value.replace(" ", "")
                if len(compact) == 5 and compact[:2] in ("SB", "SW"):
                    compact = f"{compact[:2]} {compact[2:]}"
                if not _contract_re.match(compact):
                    return jsonify({"error": f"Invalid contract '{value}' for '{field}'. Expected format 'SB K26' or 'SW K26'."}), 400
                value = compact

            # Constrain to allowed dropdown values
            if field in ALLOWED_VALUES and value not in ("", None) and value not in ALLOWED_VALUES[field]:
                return jsonify({"error": f"Invalid value '{value}' for field '{field}'"}), 400

            # Type coercion
            if field in NUMERIC_COLS:
                value = float(value) if value not in (None, "") else None

            # Safe JSON mutation via copy
            new_data = dict(deal.data or {})
            new_data[field] = value
            deal.data = new_data
            flag_modified(deal, "data")

        db.session.commit()
        return jsonify({"ok": True, "updated": len(changes)})
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@raws_bp.route("/raws/api/add", methods=["POST"])
def api_add():
    """Add a new blank deal row."""
    data = request.get_json()
    book = (data or {}).get("book", "").strip()
    if book not in ("Raws", "Whites"):
        return jsonify({"error": "book must be 'Raws' or 'Whites'"}), 400

    max_idx = db.session.query(db.func.max(PhysicalDeal.row_index)).filter_by(book=book).scalar()
    new_idx = (max_idx or 0) + 1

    cols = INPUT_COLS if book == "Raws" else WHITES_INPUT_COLS
    blank_data = {col: None for col in cols}

    deal = PhysicalDeal(book=book, row_index=new_idx, data=blank_data)
    db.session.add(deal)
    db.session.commit()
    return jsonify({"ok": True, "id": deal.id})


@raws_bp.route("/raws/api/delete", methods=["POST"])
def api_delete():
    """Delete deal rows by IDs."""
    data = request.get_json()
    ids = (data or {}).get("ids", [])
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "ids must be a non-empty list"}), 400

    deleted = PhysicalDeal.query.filter(PhysicalDeal.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True, "deleted": deleted})


@raws_bp.route("/raws/api/highlight", methods=["POST"])
def api_highlight():
    """Set or clear the row highlight color on a PhysicalDeal.

    Stored inside ``data["_highlight"]`` (shared, not per-user).
    Wiped on Excel reseed/upload since those rebuild ``data`` from the sheet.
    """
    from sqlalchemy.orm.attributes import flag_modified
    body = request.get_json() or {}
    record_id = body.get("record_id")
    color = body.get("color")
    if color not in (None, "", "yellow", "green", "red", "blue"):
        return jsonify({"error": "invalid color"}), 400
    deal = db.session.get(PhysicalDeal, record_id)
    if not deal:
        return jsonify({"error": "not found"}), 404
    new_data = dict(deal.data or {})
    if not color:
        new_data.pop("_highlight", None)
    else:
        new_data["_highlight"] = color
    deal.data = new_data
    flag_modified(deal, "data")
    db.session.commit()
    return jsonify({"ok": True})
