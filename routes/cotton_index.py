"""Cotton Index page — read-only display of the three reference tables that
live on the workbook's `Index` sheet (`Index_Basis`, `Index_Colour`,
`Index_Staple`). Data is sourced via two paths:

  - "Seed from Disk" reads the canonical workbook at EXCEL_PATH (project root).
  - "Override from Excel" accepts an uploaded .xlsm/.xlsx for ad-hoc replace.

Both paths replace all `cotton_index_rows` for the three known tables in a
single transaction. No inline editing, no PnL integration in this version.
"""
import logging
import re
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from flask import Blueprint, jsonify, render_template, request

from models.cotton import CottonIndexRow
from models.db import db

cotton_index_bp = Blueprint("cotton_index", __name__)
logger = logging.getLogger(__name__)

# Canonical workbook lives in the project root with a version suffix
# (cottonm2m28.xlsm, cottonm2m29.xlsm, ...). The reference copy lives at
# Documents\Reconciliation\cottonm2m{N}.xlsm; copy the latest into the
# project root on each refresh — no rename needed. _resolve_excel_path()
# picks the highest-numbered file at fetch time so version bumps don't
# require a code change.
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_GLOB = "cottonm2m*.xlsm"
_VERSION_RE = re.compile(r"cottonm2m(\d+)\.xlsm$", re.IGNORECASE)


def _resolve_excel_path():
    """Return the project-root cottonm2m*.xlsm with the highest version
    number, or None if no matching file exists. Files without a numeric
    suffix sort lowest so they only win if nothing versioned is present."""
    candidates = list(PROJECT_ROOT.glob(EXCEL_GLOB))
    if not candidates:
        return None

    def _version(p):
        m = _VERSION_RE.search(p.name)
        return int(m.group(1)) if m else -1

    return max(candidates, key=_version)


INDEX_TABLES = ("Index_Basis", "Index_Colour", "Index_Staple")


def _json_safe(v):
    """Normalize Python values that db.JSON (JSONB) can't serialize natively."""
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    return v


def _parse_index_workbook(wb):
    """Parse the Index sheet's three Excel tables.

    Returns {'Index_Basis': {'headers': [...], 'rows': [(row_index, data), ...]}, ...}.
    Headers come straight from each table ref's first row; data rows are dicts
    built via dict(zip(headers, values)) so all rows for a given table share the
    same key order. Skips fully-blank rows. Raises ValueError if the Index
    sheet or any required table is missing."""
    if "Index" not in wb.sheetnames:
        raise ValueError("Workbook is missing the 'Index' sheet")
    ws = wb["Index"]
    missing = [t for t in INDEX_TABLES if t not in ws.tables]
    if missing:
        raise ValueError(f"Index sheet is missing required tables: {', '.join(missing)}")

    out = {}
    for tname in INDEX_TABLES:
        rows = list(ws[ws.tables[tname].ref])
        if not rows:
            out[tname] = {"headers": [], "rows": []}
            continue
        headers = [
            str(c.value).strip() if c.value is not None else f"col_{i}"
            for i, c in enumerate(rows[0])
        ]
        parsed = []
        for i, row in enumerate(rows[1:]):
            values = [_json_safe(c.value) for c in row]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            parsed.append((i, dict(zip(headers, values))))
        out[tname] = {"headers": headers, "rows": parsed}
    return out


def _replace_rows(parsed, source):
    """Replace all CottonIndexRow rows for the three known tables in one txn."""
    CottonIndexRow.query.filter(
        CottonIndexRow.table_name.in_(INDEX_TABLES)
    ).delete(synchronize_session=False)
    for tname, payload in parsed.items():
        for row_index, data in payload["rows"]:
            db.session.add(CottonIndexRow(
                table_name=tname,
                row_index=row_index,
                data=data,
                source=source,
            ))
    db.session.commit()


def _load_table(table_name):
    """Return (headers, rows) for a single table_name, in row_index order.

    Each row is `{"id": <int>, "data": {<header>: <value>, ...}}`. Headers come
    from the first row's data keys — the parser stores all rows with the same
    key order, and both Python dicts (3.7+) and Postgres JSONB preserve
    insertion order, so this is the Excel header order."""
    rows = (CottonIndexRow.query
            .filter_by(table_name=table_name)
            .order_by(CottonIndexRow.row_index)
            .all())
    if not rows:
        return [], []
    headers = list(rows[0].data.keys())
    return headers, [{"id": r.id, "data": r.data} for r in rows]


# Index Basis: derived columns are recomputed from FOB Basis at render time.
# The DB still holds whatever the Excel sheet had so the seed/upload roundtrip
# stays faithful, but the displayed values are authoritative.
BASIS_CALCULATED_COLS = ("Bd", "India", "Staple 1 3/16")


def _round2(v):
    return round(v, 2)


def _basis_calculated_values(data):
    """Return {col: value} for the three Index_Basis calculated columns,
    given a row's data dict. None for any column that can't be computed."""
    fob = data.get("FOB Basis")
    if not isinstance(fob, (int, float)):
        return {col: None for col in BASIS_CALCULATED_COLS}
    bd = _round2(fob + 3.2)
    return {
        "Bd":            bd,
        "India":         _round2(fob + 2.8),
        "Staple 1 3/16": _round2(bd + 1.6),
    }


def _apply_basis_formulas(rows):
    """Overwrite Bd / India / Staple 1 3/16 in each row's data from FOB Basis.
    Returns a new list of {"id", "data"} dicts; originals are not mutated."""
    out = []
    for r in rows:
        new_data = dict(r["data"])
        new_data.update(_basis_calculated_values(new_data))
        out.append({"id": r["id"], "data": new_data})
    return out


@cotton_index_bp.route("/index")
def index():
    basis_headers,  basis_rows  = _load_table("Index_Basis")
    colour_headers, colour_rows = _load_table("Index_Colour")
    staple_headers, staple_rows = _load_table("Index_Staple")
    basis_rows = _apply_basis_formulas(basis_rows)
    return render_template(
        "cotton/index.html",
        basis_headers=basis_headers,   basis_rows=basis_rows,
        colour_headers=colour_headers, colour_rows=colour_rows,
        staple_headers=staple_headers, staple_rows=staple_rows,
        basis_calculated_cols=BASIS_CALCULATED_COLS,
    )


@cotton_index_bp.route("/index/api/seed", methods=["POST"])
def api_seed():
    """Seed cotton_index_rows from the highest-versioned cottonm2m*.xlsm in the project root."""
    excel_path = _resolve_excel_path()
    if excel_path is None:
        return jsonify({
            "error": f"No {EXCEL_GLOB} found in {PROJECT_ROOT}",
        }), 422
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)
        try:
            parsed = _parse_index_workbook(wb)
        finally:
            wb.close()
        _replace_rows(parsed, source="seed")
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Cotton index seed failed")
        return jsonify({"error": str(e)}), 500
    counts = {tname: len(parsed[tname]["rows"]) for tname in INDEX_TABLES}
    logger.info("Cotton index seed from %s: %s", excel_path.name, counts)
    return jsonify({
        "ok": True,
        "source_file": excel_path.name,
        "basis":  counts["Index_Basis"],
        "colour": counts["Index_Colour"],
        "staple": counts["Index_Staple"],
    })


def _coerce_cell_value(v):
    """Empty strings -> None; whole-number strings -> int; numeric strings ->
    float; everything else stays as string."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return s


@cotton_index_bp.route("/index/api/row/<int:row_id>", methods=["POST"])
def api_update_row(row_id):
    """Update a single editable cell in cotton_index_rows. Calculated columns
    are server-side rejected (defense in depth — UI also hides the affordance).
    If FOB Basis on Index_Basis is edited, response includes the recomputed
    calculated columns so the client can update them in place."""
    row = CottonIndexRow.query.get_or_404(row_id)
    body = request.get_json(silent=True) or {}
    column = body.get("column")
    if not column or column not in row.data:
        return jsonify({"error": f"Unknown column '{column}'"}), 400
    if row.table_name == "Index_Basis" and column in BASIS_CALCULATED_COLS:
        return jsonify({"error": f"Column '{column}' is calculated and cannot be edited"}), 400

    coerced = _coerce_cell_value(body.get("value"))
    new_data = dict(row.data)
    new_data[column] = coerced
    row.data = new_data  # reassign so SQLAlchemy detects the JSON mutation
    db.session.commit()

    response = {"ok": True, "value": coerced}
    if row.table_name == "Index_Basis" and column == "FOB Basis":
        response["calculated"] = _basis_calculated_values(new_data)
    return jsonify(response)


@cotton_index_bp.route("/index/api/upload", methods=["POST"])
def api_upload():
    """Upload an .xlsm / .xlsx and replace cotton_index_rows from its Index sheet."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        wb = openpyxl.load_workbook(f, data_only=True, keep_vba=True)
        try:
            parsed = _parse_index_workbook(wb)
        finally:
            wb.close()
        _replace_rows(parsed, source="upload")
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Cotton index upload failed")
        return jsonify({"error": str(e)}), 500
    counts = {tname: len(parsed[tname]["rows"]) for tname in INDEX_TABLES}
    logger.info("Cotton index upload: %s", counts)
    return jsonify({
        "ok": True,
        "basis":  counts["Index_Basis"],
        "colour": counts["Index_Colour"],
        "staple": counts["Index_Staple"],
    })
