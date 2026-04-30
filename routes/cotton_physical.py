"""Cotton Physical Deals — /cotton/physical (Purchases tab; Sales is Phase 2 placeholder).

Mirrors sugar /raws structurally but stripped down: input columns from cottonm2m.xlsm only,
no calculated columns yet. JSON-blob storage in cotton_physical_deals lets us add columns
later without a migration.
"""
import logging
import os
from datetime import datetime, date
from pathlib import Path
from flask import Blueprint, render_template, request, jsonify
import openpyxl

from models.db import db
from models.cotton import CottonPhysicalDeal

logger = logging.getLogger(__name__)

cotton_physical_bp = Blueprint("cotton_physical", __name__)

EXCEL_PATH = Path(__file__).parent.parent / "cottonm2m.xlsm"

# When true (default), an empty Purchases table auto-seeds from cottonm2m.xlsm
# on the first page load. Set EXCEL_AUTOSEED=false on Azure to disable; the
# explicit "Sync from Excel" / "Override from Excel" buttons keep working.
EXCEL_AUTOSEED = os.environ.get("EXCEL_AUTOSEED", "true").lower() != "false"

# Authoritative list of Phase 1 input columns for the Purchases sheet.
# Matches Excel headers in row 2 of cottonm2m.xlsm "Purchases" sheet.
# Calculated columns (Pricing, Quotes, Currency, PNL groups, plus computed fields
# in Additional Info) are intentionally excluded — added one by one in later phases.
PURCHASES_INPUT_COLS = [
    # Contractual Inputs (18)
    "Entry Date", "Counterparty", "Contract Reference", "Volume", "Status",
    "Origin", "Year", "Colour", "Staple", "Start Shipment", "End Shipment",
    "Terms", "Term Input", "Currency", "Exchange Rate", "Forward Rate",
    "Additional Costs", "Original Contract",
    # Trading Inputs (2)
    "Execution Contract", "Cycle Time",
    # Additional Info (1 — only the manual override column)
    "Exceptions",
]

SALES_INPUT_COLS = []  # Phase 2.

# Calculated columns — read-only, derived in index() from other tables.
PURCHASES_CALCULATED_COLS = [
    "Original Contract Settlement",
    "Futures Pricing",
    "FOB Price",
    "Volume in Lots",
    "Tagged Lots",
    "Hedge %",
    "Hedge Status",
]

# Calc columns rendered as percentages (value × 100, with % suffix).
PURCHASES_PERCENT_COLS = {"Hedge %"}

# Calc columns rendered at 1 decimal place (others default to 2 dp).
PURCHASES_ONE_DP_COLS = {"Volume in Lots", "Tagged Lots"}

# Display order: input cols + calculated cols, in the order they appear on the page.
PURCHASES_DISPLAY_COLS = [
    # Contractual Inputs
    "Entry Date", "Counterparty", "Contract Reference", "Volume", "Status",
    "Origin", "Year", "Colour", "Staple", "Start Shipment", "End Shipment",
    "Terms", "Term Input", "Currency", "Exchange Rate", "Forward Rate",
    "Additional Costs", "Original Contract",
    # Trading Inputs
    "Execution Contract", "Cycle Time",
    # Pricing
    "Original Contract Settlement", "Futures Pricing", "FOB Price",
    # Quantity & Hedging
    "Volume in Lots", "Tagged Lots", "Hedge %", "Hedge Status",
    # Additional Info
    "Exceptions",
]

PURCHASES_COLUMN_GROUPS = {
    "Contractual Inputs": [
        "Entry Date", "Counterparty", "Contract Reference", "Volume", "Status",
        "Origin", "Year", "Colour", "Staple", "Start Shipment", "End Shipment",
        "Terms", "Term Input", "Currency", "Exchange Rate", "Forward Rate",
        "Additional Costs", "Original Contract",
    ],
    "Trading Inputs": ["Execution Contract", "Cycle Time"],
    "Pricing": ["Original Contract Settlement", "Futures Pricing", "FOB Price"],
    "Quantity & Hedging": ["Volume in Lots", "Tagged Lots", "Hedge %", "Hedge Status"],
    "Additional Info": ["Exceptions"],
}

# Hedge % threshold above which a deal is considered fully priced/hedged.
# Mirrors the 0.93 cutoff in the Excel Hedge Status array formula.
HEDGE_FULL_THRESHOLD = 0.93

PURCHASES_GROUP_IDS = {
    "Contractual Inputs": "contractual",
    "Trading Inputs": "trading",
    "Pricing": "pricing",
    "Quantity & Hedging": "qty_hedging",
    "Additional Info": "additional",
}

PURCHASES_GROUP_COLORS = {
    "Contractual Inputs": "#cfe2ff",
    "Trading Inputs": "#fff3cd",
    "Pricing": "#fce4d6",
    "Quantity & Hedging": "#e2d9f3",
    "Additional Info": "#d1e7dd",
}

# 1 cotton lot = 22.7 metric tons. Used in Volume in Lots and Hedge % calculations.
COTTON_MT_PER_LOT = 22.7

NUMERIC_COLS = {
    "Volume", "Term Input", "Exchange Rate", "Forward Rate",
    "Additional Costs", "Cycle Time",
}

DATE_COLS = {"Entry Date", "Start Shipment", "End Shipment"}


def _normalize(value):
    """Normalize Excel cell values for JSON storage and consistent rendering.

    datetime/date → 'YYYY-MM-DD' string; numbers passed through (NaN → None);
    strings stripped (empty → None); anything else stringified.
    """
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value != value:  # NaN check
            return None
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        s = value.strip()
        return s if s else None
    return str(value)


def _parse_workbook_purchases(wb):
    """Parse the Purchases sheet of cottonm2m.xlsm into payloads.

    Returns (payloads, duplicates) where:
      payloads — list of {"book": "Purchases", "row_index": i, "data": {...}, "source": "..."}
      duplicates — list of Contract References that appear more than once in the file.

    Header-based parsing: row 1 is the group header (skipped), row 2 is column headers,
    row 3+ is data. Survives column inserts/reorders in Excel.
    """
    if "Purchases" not in wb.sheetnames:
        raise ValueError("cottonm2m.xlsm is missing the 'Purchases' sheet")
    ws = wb["Purchases"]

    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    headers = {h: idx for idx, h in enumerate(header_row) if h}
    missing = [c for c in PURCHASES_INPUT_COLS if c not in headers]
    if missing:
        raise ValueError(f"Purchases sheet missing headers: {missing}")

    payloads = []
    seen_refs = set()
    dup_refs = []
    idx = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None or v == "" for v in row):
            continue
        # Skip rows where the natural key is blank — Excel often has trailing
        # formula-only rows that openpyxl includes in dimensions.
        ref = row[headers["Contract Reference"]] if "Contract Reference" in headers else None
        if ref is None or (isinstance(ref, str) and not ref.strip()):
            continue
        data = {col: _normalize(row[headers[col]]) for col in PURCHASES_INPUT_COLS}
        ref_key = data.get("Contract Reference")
        if ref_key in seen_refs:
            dup_refs.append(ref_key)
        else:
            seen_refs.add(ref_key)
        payloads.append({"book": "Purchases", "row_index": idx, "data": data})
        idx += 1
    return payloads, dup_refs


def _replace_purchases(payloads, source_label):
    """Atomically wipe Purchases rows and bulk-insert new payloads."""
    deleted = CottonPhysicalDeal.query.filter_by(book="Purchases").delete(synchronize_session=False)
    if payloads:
        for p in payloads:
            p["source"] = source_label
        db.session.bulk_insert_mappings(CottonPhysicalDeal, payloads)
    db.session.commit()
    return deleted


def _load_trade_aggs(book='Purchases'):
    """Return ``(tagged_lots_map, price_total_map)`` aggregated from cotton trades.

    Both dicts are keyed by ``(contract_norm, ref)`` and mirror the Excel SUMIFS
    pair (Tagged Lots and Price Total) over the same filters:
      Trades[Book]="Physical", Trades[Instrument]="Futures",
      Trades[Contract]=<row contract>, Trades[Contract Reference]=<row ref>.

    Tagged Lots = sum(Long + Short).
    Price Total = sum((Long + Short) * Price).

    For Purchases, the reference column on trades is ``New_AGP__r.Name``
    (e.g. 'AGP/24/10/24458'); for Sales it'll be ``New_AGS__r.Name``.
    """
    from models.cotton import CottonTradePosition
    ref_field = "New_AGP__r.Name" if book == "Purchases" else "New_AGS__r.Name"
    trades = CottonTradePosition.query.filter(
        CottonTradePosition.book_parsed == "Physical",
        CottonTradePosition.instrument == "Futures",
    ).all()
    tagged = {}
    price_total = {}
    for t in trades:
        d = t.data or {}
        contract = (d.get("Contract__c") or "").replace(" ", "").upper()
        ref = d.get(ref_field)
        if not contract or not ref:
            continue
        net = float(d.get("Long__c") or 0) + float(d.get("Short__c") or 0)
        price = float(d.get("Price__c") or 0)
        key = (contract, ref)
        tagged[key] = tagged.get(key, 0) + net
        price_total[key] = price_total.get(key, 0) + net * price
    return tagged, price_total


def _fob_price(terms, term_input, futures_pricing, currency, forward_rate):
    """Compute the Excel FOB Price formula:

      =IF(ISBLANK(Terms), "-",
        IF(Terms="Flat", Term Input,
        Futures Pricing + Contract Basis))
      * IF(Currency="CFA", (1/655.957) * Forward Rate * 100 / 2.20462, 1)

    For Basis terms the inner Contract Basis equals Term Input (Contract Basis's
    own Basis-branch formula), so we substitute Term Input directly here. This
    keeps FOB Price independent of the Contract Basis column.

    Currency='CFA' applies a CFA→USD/lb conversion: (1/655.957 EUR/CFA) ×
    Forward Rate (USD/EUR) × 100 (USD→cents) / 2.20462 (lb/kg).
    """
    if not terms or term_input is None:
        return None
    if terms == "Flat":
        base = term_input
    elif terms == "Basis":
        if futures_pricing is None or not isinstance(futures_pricing, (int, float)):
            return None
        base = futures_pricing + term_input
    else:
        return None
    if currency == "CFA":
        if forward_rate is None or forward_rate == 0:
            return None
        factor = (1 / 655.957) * forward_rate * 100 / 2.20462
    else:
        factor = 1
    return base * factor


def _futures_pricing(hedge_status, terms, settlement, vol_lots, tagged_lots, price_total):
    """Compute the Excel IFS Futures Pricing formula.

    Branches in the order specified by the workbook formula. Returns a number
    for priced/hedged cases, the literal string "Unhedged" for that status
    (matching the Excel branch that returns the status text), or None if
    required inputs are missing.
    """
    if hedge_status == "Unpriced":
        return settlement
    if hedge_status == "Partial":
        if None in (price_total, vol_lots, tagged_lots, settlement) or vol_lots == 0:
            return None
        return (abs(price_total) + abs(vol_lots + tagged_lots) * settlement) / vol_lots
    if hedge_status == "Priced":
        if None in (price_total, tagged_lots) or tagged_lots == 0:
            return None
        return price_total / tagged_lots
    if hedge_status == "Unhedged":
        return "Unhedged"
    if terms == "Flat":
        if None in (price_total, tagged_lots) or tagged_lots == 0:
            return None
        return price_total / tagged_lots
    return None


def _seed_purchases_from_disk():
    """Auto-seed Purchases from EXCEL_PATH if the file exists. Used as fallback."""
    if not EXCEL_PATH.exists():
        return
    logger.info("Auto-seed fallback triggered for cotton Purchases from %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        payloads, _ = _parse_workbook_purchases(wb)
    finally:
        wb.close()
    if payloads:
        _replace_purchases(payloads, source_label="excel-default")
        logger.info("Auto-seeded %d cotton Purchases rows", len(payloads))


@cotton_physical_bp.route("/physical")
def index():
    from services.price_source import load_price_map
    from models.cotton import CottonMarketPrice

    error = None
    rows = []
    try:
        # Sett-1 prices keyed by normalised contract (e.g. 'CTK25').
        # Always Sett-1 here regardless of global toggle — matches the Excel
        # formula. A future "Live Contract Settlement" column will be the
        # toggle-aware live equivalent.
        sett1_prices, _ = load_price_map(source='sett1', model=CottonMarketPrice)
        # Tagged Lots + Price Total: SUMIFS lookups against cotton trades (Physical/Futures).
        tagged_lots_map, price_total_map = _load_trade_aggs(book='Purchases')

        deals = CottonPhysicalDeal.query.filter_by(book="Purchases").order_by(CottonPhysicalDeal.row_index).all()
        if not deals and EXCEL_AUTOSEED:
            _seed_purchases_from_disk()
            deals = CottonPhysicalDeal.query.filter_by(book="Purchases").order_by(CottonPhysicalDeal.row_index).all()
        for deal in deals:
            d = deal.data or {}
            row = {col: d.get(col) for col in PURCHASES_INPUT_COLS}
            row["_deal_id"] = deal.id
            # Calculated: Original Contract Settlement = Sett-1 lookup of Original Contract
            orig = d.get("Original Contract")
            key = str(orig).replace(" ", "").upper() if orig else ""
            row["Original Contract Settlement"] = sett1_prices.get(key) if key else None
            # Calculated: Volume in Lots = Volume / 22.7
            volume = d.get("Volume")
            try:
                row["Volume in Lots"] = float(volume) / COTTON_MT_PER_LOT if volume not in (None, "") else None
            except (TypeError, ValueError):
                row["Volume in Lots"] = None
            # Calculated: Tagged Lots = SUMIFS(Trades[Position], filters)
            ref = d.get("Contract Reference")
            row["Tagged Lots"] = tagged_lots_map.get((key, ref), 0) if (key and ref) else None
            # Calculated: Hedge % = IF(Volume in Lots=0, 0, ABS(Tagged Lots)/ABS(Volume in Lots))
            vl = row["Volume in Lots"]
            tl = row["Tagged Lots"]
            if vl is None or tl is None:
                row["Hedge %"] = None
            elif vl == 0:
                row["Hedge %"] = 0
            else:
                row["Hedge %"] = abs(tl) / abs(vl)
            # Calculated: Hedge Status — branches by Terms then by Hedge % threshold.
            terms = d.get("Terms")
            hp = row["Hedge %"]
            if terms not in ("Basis", "Flat") or hp is None:
                row["Hedge Status"] = None
            elif terms == "Basis":
                if hp == 0:
                    row["Hedge Status"] = "Unpriced"
                elif hp < HEDGE_FULL_THRESHOLD:
                    row["Hedge Status"] = "Partial"
                else:
                    row["Hedge Status"] = "Priced"
            else:  # Flat
                if hp == 0:
                    row["Hedge Status"] = "Unhedged"
                elif hp < HEDGE_FULL_THRESHOLD:
                    row["Hedge Status"] = "Exposure"
                else:
                    row["Hedge Status"] = "Hedged"
            # Calculated: Futures Pricing — IFS by Hedge Status / Terms.
            row["Futures Pricing"] = _futures_pricing(
                row["Hedge Status"],
                terms,
                row["Original Contract Settlement"],
                row["Volume in Lots"],
                row["Tagged Lots"],
                price_total_map.get((key, ref)) if (key and ref) else None,
            )
            # Calculated: FOB Price — Term Input (Flat) or Futures Pricing+Term Input (Basis),
            # times CFA→USD/lb conversion if Currency='CFA'.
            row["FOB Price"] = _fob_price(
                terms,
                d.get("Term Input"),
                row["Futures Pricing"],
                d.get("Currency"),
                d.get("Forward Rate"),
            )
            rows.append(row)
    except Exception as e:
        logger.exception("Failed to load cotton Purchases")
        error = str(e)

    return render_template(
        "cotton/physical.html",
        rows=rows,
        columns=PURCHASES_DISPLAY_COLS,
        calculated_cols=set(PURCHASES_CALCULATED_COLS),
        percent_cols=PURCHASES_PERCENT_COLS,
        one_dp_cols=PURCHASES_ONE_DP_COLS,
        column_groups=PURCHASES_COLUMN_GROUPS,
        group_ids=PURCHASES_GROUP_IDS,
        group_colors=PURCHASES_GROUP_COLORS,
        numeric_cols=NUMERIC_COLS,
        date_cols=DATE_COLS,
        error=error,
    )


@cotton_physical_bp.route("/physical/api/seed", methods=["POST"])
def api_seed():
    """Seed Purchases from on-disk cottonm2m.xlsm. Replaces all current Purchases rows."""
    if not EXCEL_PATH.exists():
        return jsonify({"error": f"File not found: {EXCEL_PATH}"}), 400
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        try:
            payloads, dup_refs = _parse_workbook_purchases(wb)
        finally:
            wb.close()
        deleted = _replace_purchases(payloads, source_label="excel-default")
        logger.info("Cotton Purchases seed: deleted=%d inserted=%d duplicates=%d",
                    deleted, len(payloads), len(dup_refs))
        return jsonify({"ok": True, "deleted": deleted, "inserted": len(payloads),
                        "duplicates": dup_refs})
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Cotton Purchases seed failed")
        return jsonify({"error": str(e)}), 500


@cotton_physical_bp.route("/physical/api/upload", methods=["POST"])
def api_upload():
    """Upload an .xlsm to replace Purchases. Emergency override."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            payloads, dup_refs = _parse_workbook_purchases(wb)
        finally:
            wb.close()
        deleted = _replace_purchases(payloads, source_label="excel-upload")
        logger.info("Cotton Purchases upload: deleted=%d inserted=%d duplicates=%d",
                    deleted, len(payloads), len(dup_refs))
        return jsonify({"ok": True, "deleted": deleted, "inserted": len(payloads),
                        "duplicates": dup_refs})
    except ValueError as e:
        return jsonify({"error": str(e)}), 422
    except Exception as e:
        db.session.rollback()
        logger.exception("Cotton Purchases upload failed")
        return jsonify({"error": str(e)}), 500


@cotton_physical_bp.route("/physical/api/update", methods=["POST"])
def api_update():
    """Inline-edit one or more fields on cotton physical deals."""
    from sqlalchemy.orm.attributes import flag_modified
    body = request.get_json() or {}
    changes = body.get("changes", [])
    if not changes:
        return jsonify({"ok": True, "updated": 0})

    allowed = set(PURCHASES_INPUT_COLS)
    try:
        for change in changes:
            record_id = change.get("record_id")
            field = (change.get("field") or "").strip()
            value = change.get("value")
            if field not in allowed:
                return jsonify({"error": f"Disallowed field '{field}'"}), 400
            deal = db.session.get(CottonPhysicalDeal, record_id)
            if not deal:
                continue
            # Type coercion
            if field in NUMERIC_COLS:
                value = float(value) if value not in (None, "") else None
            elif isinstance(value, str):
                value = value.strip() or None
            new_data = dict(deal.data or {})
            new_data[field] = value
            deal.data = new_data
            deal.source = "manual-edit"
            flag_modified(deal, "data")
        db.session.commit()
        return jsonify({"ok": True, "updated": len(changes)})
    except Exception as e:
        db.session.rollback()
        logger.exception("Cotton Purchases update failed")
        return jsonify({"error": str(e)}), 500


@cotton_physical_bp.route("/physical/api/delete", methods=["POST"])
def api_delete():
    """Delete deal rows by IDs."""
    body = request.get_json() or {}
    ids = body.get("ids", [])
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "ids must be a non-empty list"}), 400
    deleted = CottonPhysicalDeal.query.filter(CottonPhysicalDeal.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True, "deleted": deleted})
